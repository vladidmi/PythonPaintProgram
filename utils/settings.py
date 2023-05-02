import os
import datetime
import holidays
from pandas.tseries.offsets import CustomBusinessDay
from PIL import Image, ImageDraw, ImageFont
import pandas as pd
import numpy as np
import re
import plotly.express as px
import openpyxl


def hex_colour_to_rgb(hex_colour):
    current_r = int(str(hex_colour[-6:-4]), base=16)
    current_g = int(str(hex_colour[-4:-2]), base=16)
    current_b = int(str(hex_colour[-2:]), base=16)
    return (current_r, current_g, current_b)


def resize_image(image_path):
    # resizing the image to fit into the given rectangular
    image = Image.open(image_path)
    current_image_width, current_image_height = image.size
    resize_ratio = min(WIDTH / current_image_width, HEIGHT / current_image_height)
    new_image_width = int(current_image_width * resize_ratio)
    new_image_height = int(current_image_height * resize_ratio)
    image_resized = image.resize((new_image_width, new_image_height))
    blank_image = Image.new(mode="RGB", size=(WIDTH, HEIGHT), color=WHITE)
    offset = ((WIDTH - new_image_width) // 2, (HEIGHT - new_image_height) // 2)
    blank_image.paste(image_resized, offset)
    blank_image.save(image_path)


def weekdays_of_current_week(current_date, project_info):
    current_date = list(current_date.isocalendar())
    if project_info["Arbeiten am Samstag [ja/nein]"] == "ja":
        weekdays_to_print = 7
    else:
        weekdays_to_print = 6

    weekdays = []
    for i in range(1, weekdays_to_print):
        current_date[2] = i
        current_date_iso = datetime.datetime.fromisocalendar(*current_date).date()
        if current_date_iso not in german_holidays:
            weekdays.append(current_date_iso)

    return weekdays


def translate_days_to_german(day):
    for german_week_day in GERMAN_WEEK_DAYS:
        day = day.replace(*german_week_day)
    return day


class Floor_level_info:
    def __init__(self, image_name, full_image_name, floor_id):
        self.image_name = image_name
        self.full_path_image = full_image_name
        self.full_path_xlsx = full_image_name.replace("jpg", "xlsx")
        self.floor_name = image_name.replace(".jpg", "")
        self.floor_id = floor_id
        self.image = Image.open(self.full_path_image)
        self.image_width, self.image_height = self.image.size


cur_dir = os.getcwd()

test_file = os.path.join(cur_dir, "utils", "projekt_info.xlsx")


def read_file_for_working_steps(filename):
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)
    # Get the specific worksheet
    worksheet = workbook["Arbeitsschritte"]
    # Initialize the dictionary to store the data
    data = {}
    # Iterate through the rows
    for row in worksheet.iter_rows():
        # Get the key from the first column
        key = row[0].value
        if key == "Strukturtyp":
            continue
        elif key == None:
            break
        # Initialize the dictionary for the current row
        row_data = {}
        # Iterate through the remaining columns
        for cell in row[1:]:
            # Get the value and color of the cell
            value = cell.value
            if value == None:
                continue
            else:
                value = f"{key}@{value}"
            color = str(cell.fill.start_color.index)
            # Add the data to the dictionary for the current row
            row_data[value] = "#" + color.lower()[2:]
        # Add the data for the current row to the main dictionary
        data[key] = {}
        data[key]["Arbeitsschritte"] = row_data
        data[key]["Strukturtypfarbe"] = (
            "#" + str(row[0].fill.start_color.index).lower()[2:]
        )
    # Return the dictionary
    return data


def read_file_for_project_info(filename):
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)
    # Get the specific worksheet
    worksheet = workbook["Projekt"]
    # Initialize the dictionary to store the data
    data = {}
    # Iterate through the rows
    for row in worksheet.iter_rows():
        # Get the key from the first column
        key = row[0].value
        # Initialize the dictionary for the current row
        row_data = {}
        # Add the data for the current row to the main dictionary
        data[key] = row[1].value
    # Return the dictionary
    return data


def read_file_for_tact(filename):
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)
    # Get the specific worksheet
    worksheet = workbook["Bauabschnitte"]
    # Initialize the dictionary to store the data
    data = []
    # Iterate through the rows
    for row in worksheet.iter_rows():
        # Get the key from the first column
        key = row[0].value
        if key == None:
            break
        color = "#" + row[0].fill.start_color.index[2:]
        if color == "#":
            print(f"problem with {row[0].value}")
        data.append({"tact_text": key, "tact_color": color.lower()})
    # Return the list with tacts
    return data


working_steps = read_file_for_working_steps(test_file)

project_info = read_file_for_project_info(test_file)

tact_info = read_file_for_tact(test_file)

print({key: working_steps[key]["Strukturtypfarbe"] for key in working_steps})
print(working_steps)
print(project_info)

for i in tact_info:
    print(i)

MAX_GRID_SIZE = 400

BUTTON_TEXT_SIZE = 8
PROJECT_INFO_TEXT_SIZE = 14

project_font_path = ImageFont.truetype(
    os.path.join(cur_dir, "utils", "arial.ttf"), PROJECT_INFO_TEXT_SIZE
)
project_font_path_small = ImageFont.truetype(
    os.path.join(cur_dir, "utils", "arial.ttf"), BUTTON_TEXT_SIZE
)

legend_font_path_medium = ImageFont.truetype(
    os.path.join(cur_dir, "utils", "arial.ttf"), int(BUTTON_TEXT_SIZE * 1.5)
)

path_to_image_folder = os.path.join(cur_dir, "imgs")
path_to_temp_image = os.path.join(cur_dir, "utils", "temp.jpg")
full_image_path = {
    file: os.path.join(path_to_image_folder, file)
    for file in os.listdir(path_to_image_folder)
    if ".jpg" in file
}

path_to_comments = os.path.join(path_to_image_folder, "project_comments.xlsx")

weekmask_germany = "Mon Tue Wed Thu Fri"
if project_info["Arbeiten am Samstag [ja/nein]"] == "ja":
    weekmask_germany += " Sat"
german_holidays = [
    date[0]
    for date in holidays.Germany(
        years=list(range(2020, 2040)), prov=project_info["Bundesland"]
    ).items()
]

for year in range(2020, 2030):
    # adding the winter pause (CW52,CW1)
    for week in [52, 1]:
        for day in range(7):
            week1 = f"{year}-W{week}"
            day1 = datetime.datetime.strptime(
                week1 + "-1", "%Y-W%W-%w"
            ) + datetime.timedelta(days=day)
            german_holidays.append(day1)

german_business_day = CustomBusinessDay(
    holidays=german_holidays, weekmask=weekmask_germany
)

WHITE = "#ffffff"
BLACK = "#000000"
RED = "#ff0000"
BLUE = "#0000ff"
GREEN = "#00ff00"
YELLOW = "#ffff00"  # Schalen
LIGHT_BLUE = "#00b0f0"  # Bewehren
LIGHT_GREEN = "#00c000"  # Betonieren
VIOLET = "#8057ff"  # Fertigteil setzen
ORANGE = "#ffc000"  # Mauern
DARK_RED = "#c00000"  # Erledigt
GREY = "#808080"
BROWN = "#644624"
PINK = "#ffc0cb"
BEIGE = "#d1bc8a"  # Leerohre

BG_COLOR = WHITE
TRANSPARENT = 0.4
SEMI_TRANSPARENT = 0.75
NOT_TRANSPARENT = 1

PHOTO_IMAGE = "photo_image"
PHOTO_IMAGE_WIDTH = "photo_image_width"
PHOTO_IMAGE_HEIGHT = "photo_image_height"
PHOTO_IMAGE_FILL = "photo_image_fill"

pixel_size_increase = 2

GERMAN_WEEK_DAYS = (
    ("Monday", "Montag"),
    ("Tuesday", "Dienstag"),
    ("Wednesday", "Mittwoch"),
    ("Thursday", "Donnerstag"),
    ("Friday", "Freitag"),
    ("Saturday", "Samstag"),
    ("Sunday", "Sonntag"),
)

HORIZONTAL = "horizontal"
VERTICAL = "vertical"
CURSOR_SIZE = "Kursorgröße"

# COMMON
ERASE = "Planung entf."
SAVE = "Speichern"
BIGGER = "Größer"
SMALLER = "Kleiner"
DRAW_MODE = "Modus"
NEXT_FLOOR = "Ebene +"
PREVIOUS_FLOOR = "Ebene -"
PRINT = "Drucken"
BUTTON_COLOUR = "Button_colour"
BUTTON_TEXT_COLOUR = "Button_text_colour"

# Drawing structure
DRAW_SCTRUCTURE = "Draw structure"
CONCRETE = "STB"
PREFABRICATED_PART = (
    "HFT."  # with point, as the name should be different from DO_MASONRY
)
MASONRY = "MW."  # with point, as the name should be different from PREFABRICATED_PART_ASSEMBLE
GROUND = "Erde."

# Planning
PLAN = "Plan"
FORMWORK = "SCH"
FORMWORK_LONG = "Schalen"
REINFORCE = "BEW"
REINFORCE_LONG = "Bewehren"
EMPTY_PIPES = "Leerrohr"
BUILT_IN_PART = "BST"
POUR_CONCRETE = "BET"
POUR_CONCRETE_LONG = "Betonieren"
PREFABRICATED_PART_ASSEMBLE = "HFT"
DO_MASONRY = "MW"
PART_COMPLETE = "Erledigt"
LAST_DAY = "Tag -"
NEXT_DAY = "Tag +"
DRAW_TEXT_ON_CANVAS = "Text einf."
DELETE_TEXT_ON_CANVAS = "Text lösch."
NEW_EVENT = "Sonst."
GROUND_JOB = "Erdarb."

long_names_for_legend = {
    FORMWORK: FORMWORK_LONG,
    REINFORCE: REINFORCE_LONG,
    POUR_CONCRETE: POUR_CONCRETE_LONG,
}

# Tact division
TACT = "BA"
TACT_LONG = "Bauabschnitte"
tact_id = None
TACT_PART = "BA."
tact_add = "BA+"
tact_delete = "BA-"
NO_TACT = "Kein BA"


# Color_dict
all_colors = {
    **{key: working_steps[key]["Strukturtypfarbe"] for key in working_steps},
    **{str(i["tact_text"]): i["tact_color"] for i in tact_info},
    NEW_EVENT: GREY,
}

print(all_colors)

for key in working_steps:
    for step_key in working_steps[key]["Arbeitsschritte"]:
        all_colors[step_key] = working_steps[key]["Arbeitsschritte"][step_key]


color_map_for_plotly = {
    step: "rgb" + str(hex_colour_to_rgb(all_colors[step])) for step in all_colors
}
