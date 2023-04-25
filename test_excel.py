import openpyxl

test_file = r"D:\projects\github\test_github\PythonPaintProgram\projekt_info.xlsx"


def read_file_for_working_steps(filename):
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)
    # Get the specific worksheet
    worksheet = workbook["Arbeitsschritte"]
    # Initialize the dictionary to store the data
    data = {}
    # Iterate through the rows
    for row in worksheet.iter_rows():
        # Get the key from the first column
        key = row[0].value
        if key == "Strukturtyp":
            continue
        # Initialize the dictionary for the current row
        row_data = {}
        # Iterate through the remaining columns
        for cell in row[1:]:
            # Get the value and color of the cell
            value = cell.value
            if value == None:
                continue
            color = cell.fill.start_color.index
            # Add the data to the dictionary for the current row
            row_data[value] = color
        # Add the data for the current row to the main dictionary
        data[key] = row_data
    # Return the dictionary
    return data


def read_file_for_project_info(filename):
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)
    # Get the specific worksheet
    worksheet = workbook["Projekt"]
    # Initialize the dictionary to store the data
    data = {}
    # Iterate through the rows
    for row in worksheet.iter_rows():
        # Get the key from the first column
        key = row[0].value
        # Initialize the dictionary for the current row
        row_data = {}
        # Add the data for the current row to the main dictionary
        data[key] = row[1].value
    # Return the dictionary
    return data


def read_file_for_tact(filename):
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)
    # Get the specific worksheet
    worksheet = workbook["Bauabschnitte"]
    # Initialize the dictionary to store the data
    data = []
    # Iterate through the rows
    for row in worksheet.iter_rows():
        # Get the key from the first column
        key = row[0].value
        color = "#" + row[0].fill.start_color.index
        data.append({key: color.lower()})
    # Return the list with tacts
    return data


working_steps = read_file_for_working_steps(test_file)

project_info = read_file_for_project_info(test_file)

tact_number = project_info.get("Anzahl der Bauabschnitte", 30)

tact_info = read_file_for_tact(test_file)

print(working_steps)
print(project_info)
print(tact_info)

for i in range(tact_number):
    print(tact_info[i])
